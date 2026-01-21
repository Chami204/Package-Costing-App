import streamlit as st
import pandas as pd
import numpy as np

# Initialize all session states for persistence
if 'primary_sku_data' not in st.session_state:
    st.session_state.primary_sku_data = pd.DataFrame(columns=[
        "SKU No", "Unit weight(kg/m)", "total weight per profile (kg)", 
        "Width/mm", "Height/mm", "Length/mm", "Box Width/mm",
        "Box Height/mm", "Box Length/mm", "W/mm", "H/mm",
        "Number of profiles per box", "Comment on fabrication"
    ])

if 'primary_material_costs' not in st.session_state:
    st.session_state.primary_material_costs = pd.DataFrame({
        "Material": ["McFoam", "Craft Paper", "Protective Tape", "Stretchwrap"],
        "Cost/ m²": [51.00, 34.65, 100.65, 14.38]
    })

if 'primary_box_costs' not in st.session_state:
    st.session_state.primary_box_costs = pd.DataFrame({
        "Length(mm)": [330], "Width (mm)": [210], 
        "Height (mm)": [135], "Cost (LKR)": [205.00]
    })


# Set page configuration
st.set_page_config(
    page_title="Packing Costing Calculator",
    page_icon="📦",
    layout="wide"
)

# Initialize session state for calculation triggers
if 'calculate_primary' not in st.session_state:
    st.session_state.calculate_primary = False
if 'calculate_secondary' not in st.session_state:
    st.session_state.calculate_secondary = False

# Function to trigger primary calculations
def trigger_primary_calculation():
    st.session_state.calculate_primary = True

# Function to trigger secondary calculations
def trigger_secondary_calculation():
    st.session_state.calculate_secondary = True

# Function to reset calculation flags
def reset_calculation_flags():
    st.session_state.calculate_primary = False
    st.session_state.calculate_secondary = False

# App title
st.title("📦 Packing Costing Calculator")

# Create download button
def create_excel_report():
    """Create Excel report with primary and secondary calculations"""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    # Create an Excel writer
    output = BytesIO()
    
    # Create border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Create heading style
    heading_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    heading_fill = PatternFill(start_color='4B0082', end_color='4B0082', fill_type='solid')  # Dark purple
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        workbook = writer.book
        
        # Sheet 1: Primary Calculations
        if 'primary_calculations' in st.session_state and not st.session_state.primary_calculations.empty:
            # Get primary selections from session state
            finish_primary = st.session_state.get("finish_primary", "Mill Finish")
            interleaving_primary = st.session_state.get("interleaving_primary", "No")
            eco_friendly_primary = st.session_state.get("eco_friendly_primary", "Mac foam")
            protective_tape_primary = st.session_state.get("protective_tape_primary", "No")
            
            # Create a summary dataframe for primary calculations
            primary_summary = pd.DataFrame({
                'Primary Calculations Summary': ['Primary Packing Costing Report']
            })
            primary_summary.to_excel(writer, sheet_name='Primary Calculations', index=False, startrow=0)
            
            # Format the primary summary
            ws = writer.sheets['Primary Calculations']
            ws['A1'].font = heading_font
            ws['A1'].fill = heading_fill
            ws['A1'].alignment = Alignment(horizontal='center')
            ws.column_dimensions['A'].width = 35
            
            # Add SKU Table
            start_row = 4
            ws.cell(row=start_row, column=1, value="SKU Table with dimensions").font = heading_font
            ws.cell(row=start_row, column=1).fill = heading_fill
            st.session_state.primary_sku_data.to_excel(writer, sheet_name='Primary Calculations', index=False, startrow=start_row + 1)
            
            # Apply borders to SKU table
            sku_data_rows = len(st.session_state.primary_sku_data)
            sku_data_cols = len(st.session_state.primary_sku_data.columns)
            for row in range(start_row + 2, start_row + sku_data_rows + 3):
                for col in range(1, sku_data_cols + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
            
            # Add Common Packing Selections heading
            common_start_row = start_row + sku_data_rows + 5
            ws.cell(row=common_start_row, column=1, value="Common Packing Selections").font = heading_font
            ws.cell(row=common_start_row, column=1).fill = heading_fill
            
            # Add Common Packing Selections details
            ws.cell(row=common_start_row + 1, column=1, value=f'Finish: {finish_primary}')
            ws.cell(row=common_start_row + 2, column=1, value=f'Interleaving Required: {interleaving_primary}')
            ws.cell(row=common_start_row + 3, column=1, value=f'Eco-Friendly Material: {eco_friendly_primary}')
            ws.cell(row=common_start_row + 4, column=1, value=f'Protective Tape: {protective_tape_primary}')
            ws.cell(row=common_start_row + 5, column=1, value=f'Box Ply: {st.session_state.get("box_ply", "3 ply")}')
            
            # Apply borders to common selections
            for row in range(common_start_row + 1, common_start_row + 6):
                cell = ws.cell(row=row, column=1)
                cell.border = thin_border
            
            # Add Material Costs heading
            material_start_row = common_start_row + 6
            ws.cell(row=material_start_row, column=1, value="Table 1: Primary Packing Material Costs").font = heading_font
            ws.cell(row=material_start_row, column=1).fill = heading_fill
            
            # Add Material Costs
            st.session_state.primary_material_costs.to_excel(writer, sheet_name='Primary Calculations', index=False, 
                                   startrow=material_start_row + 1)
            
            # Apply borders to material costs table
            material_rows = len(st.session_state.primary_material_costs)
            material_cols = len(st.session_state.primary_material_costs.columns)
            for row in range(material_start_row + 2, material_start_row + material_rows + 3):
                for col in range(1, material_cols + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
            
            # Add Cardboard Box Cost heading
            box_start_row = material_start_row + material_rows + 4
            ws.cell(row=box_start_row, column=1, value="Table 2: Cardboard Box Cost").font = heading_font
            ws.cell(row=box_start_row, column=1).fill = heading_fill
            
            # Add Cardboard Box Costs
            st.session_state.primary_box_costs.to_excel(writer, sheet_name='Primary Calculations', index=False,
                              startrow=box_start_row + 1)
            
            # Apply borders to box costs table
            box_rows = len(st.session_state.primary_box_costs)
            box_cols = len(st.session_state.primary_box_costs.columns)
            for row in range(box_start_row + 2, box_start_row + box_rows + 3):
                for col in range(1, box_cols + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
            
            # Add Primary Packing Total Cost heading
            calc_start_row = box_start_row + box_rows + 4
            ws.cell(row=calc_start_row, column=1, value="Table 3: Primary Packing Total Cost").font = heading_font
            ws.cell(row=calc_start_row, column=1).fill = heading_fill
            
            # Add Primary Packing Total Cost
            st.session_state.primary_calculations.to_excel(writer, sheet_name='Primary Calculations', index=False,
                                                            startrow=calc_start_row + 1)
            
            # Apply borders to calculations table
            calc_rows = len(st.session_state.primary_calculations)
            calc_cols = len(st.session_state.primary_calculations.columns)
            for row in range(calc_start_row + 2, calc_start_row + calc_rows + 3):
                for col in range(1, calc_cols + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
            
            # Add Special Comments Section (AFTER the table)
            comments_start_row = calc_start_row + calc_rows + 5
            ws.cell(row=comments_start_row, column=1, value="Special Comments Section").font = heading_font
            ws.cell(row=comments_start_row, column=1).fill = heading_fill
            
            # Get the comments from session state
            eco_friendly_primary = st.session_state.get("eco_friendly_primary", "Mac foam")
            comments_text = f"""Costing is done according to primary packing. Therefore, this cost does not include any crate or palletizing charges. Please note that secondary packaging will incur an additional charge.
            
            The interleaving material is "{eco_friendly_primary}".
            
            Protective tape required to avoid rejects
            
            Costing is only inclusive of interleaving required & Cardboard Box/Polybag."""
            
            # Add comments text
            current_row = comments_start_row + 2
            for line in comments_text.split('\n'):
                ws.cell(row=current_row, column=1, value=line.strip())
                current_row += 1

    
        # Sheet 2: Secondary Calculations
        if 'secondary_calculations' in st.session_state and not st.session_state.secondary_calculations.empty:
            # Get secondary selections from session state
            finish_secondary = st.session_state.get("finish_secondary", "Mill Finish")
            interleaving_secondary = st.session_state.get("interleaving_secondary", "No")
            eco_friendly_secondary = st.session_state.get("eco_friendly_secondary", "Mac foam")
            protective_tape_secondary = st.session_state.get("protective_tape_secondary", "No")
            
            # Create a summary dataframe for secondary calculations
            secondary_summary = pd.DataFrame({
                'Secondary Calculations Summary': ['Secondary Packing Costing Report']
            })
            secondary_summary.to_excel(writer, sheet_name='Secondary Calculations', index=False, startrow=0)
            
            # Format the secondary summary
            ws2 = writer.sheets['Secondary Calculations']
            ws2['A1'].font = heading_font
            ws2['A1'].fill = heading_fill
            ws2['A1'].alignment = Alignment(horizontal='center')
            ws2.column_dimensions['A'].width = 35
            
            # Add SKU Table heading
            start_row_sec = 4
            ws2.cell(row=start_row_sec, column=1, value="SKU Table with dimensions").font = heading_font
            ws2.cell(row=start_row_sec, column=1).fill = heading_fill
            
            # Add SKU Table
            st.session_state.secondary_sku_data.to_excel(writer, sheet_name='Secondary Calculations', index=False, startrow=start_row_sec + 1)
            
            # Apply borders to SKU table
            sku_data_rows_sec = len(st.session_state.secondary_sku_data)
            sku_data_cols_sec = len(st.session_state.secondary_sku_data.columns)
            for row in range(start_row_sec + 2, start_row_sec + sku_data_rows_sec + 3):
                for col in range(1, sku_data_cols_sec + 1):
                    cell = ws2.cell(row=row, column=col)
                    cell.border = thin_border
            
            # Add Common Packing Selections heading
            common_start_row_sec = start_row_sec + sku_data_rows_sec + 5
            ws2.cell(row=common_start_row_sec, column=1, value="Common Packing Selections").font = heading_font
            ws2.cell(row=common_start_row_sec, column=1).fill = heading_fill
            
            # Add Common Packing Selections details
            ws2.cell(row=common_start_row_sec + 1, column=1, value=f'Finish: {finish_secondary}')
            ws2.cell(row=common_start_row_sec + 2, column=1, value=f'Interleaving Required: {interleaving_secondary}')
            ws2.cell(row=common_start_row_sec + 3, column=1, value=f'Eco-Friendly Material: {eco_friendly_secondary}')
            ws2.cell(row=common_start_row_sec + 4, column=1, value=f'Protective Tape: {protective_tape_secondary}')
            
            # Apply borders to common selections
            for row in range(common_start_row_sec + 1, common_start_row_sec + 5):
                cell = ws2.cell(row=row, column=1)
                cell.border = thin_border
            
            # Add Bundling Data if available
            current_row = common_start_row_sec + 6
            if 'bundling_data' in st.session_state and not st.session_state.bundling_data.empty:
                ws2.cell(row=current_row, column=1, value="Section 1 - Number of layers (Method 1)").font = heading_font
                ws2.cell(row=current_row, column=1).fill = heading_fill
                
                st.session_state.bundling_data.to_excel(writer, sheet_name='Secondary Calculations', index=False,
                                                       startrow=current_row + 1, 
                                                       header=True)
                
                # Apply borders to bundling data
                bundling_rows = len(st.session_state.bundling_data)
                bundling_cols = len(st.session_state.bundling_data.columns)
                for row in range(current_row + 2, current_row + bundling_rows + 3):
                    for col in range(1, bundling_cols + 1):
                        cell = ws2.cell(row=row, column=col)
                        cell.border = thin_border
                
                current_row += bundling_rows + 5
            
            # Add Bundle Size Data if available
            if 'bundle_size_data' in st.session_state and not st.session_state.bundle_size_data.empty:
                ws2.cell(row=current_row, column=1, value="Section 2 - Size of bundle (Method 2)").font = heading_font
                ws2.cell(row=current_row, column=1).fill = heading_fill
                
                st.session_state.bundle_size_data.to_excel(writer, sheet_name='Secondary Calculations', index=False,
                                                          startrow=current_row + 1,
                                                          header=True)
                
                # Apply borders to bundle size data
                bundle_size_rows = len(st.session_state.bundle_size_data)
                bundle_size_cols = len(st.session_state.bundle_size_data.columns)
                for row in range(current_row + 2, current_row + bundle_size_rows + 3):
                    for col in range(1, bundle_size_cols + 1):
                        cell = ws2.cell(row=row, column=col)
                        cell.border = thin_border
                
                current_row += bundle_size_rows + 5
            
            # Add Cost Tables heading
            current_row += 2
            ws2.cell(row=current_row, column=1, value="Secondary Packing Cost Tables").font = heading_font
            ws2.cell(row=current_row, column=1).fill = heading_fill
            current_row += 2
            
            # Add Material Costs
            if 'secondary_material_costs' in st.session_state and not st.session_state.secondary_material_costs.empty:
                ws2.cell(row=current_row, column=1, value="Table 1: Primary Packing Material Costs").font = heading_font
                ws2.cell(row=current_row, column=1).fill = heading_fill
                
                st.session_state.secondary_material_costs.to_excel(writer, sheet_name='Secondary Calculations', index=False,
                                                       startrow=current_row + 1)
                
                # Apply borders
                mat_rows = len(st.session_state.secondary_material_costs)
                mat_cols = len(st.session_state.secondary_material_costs.columns)
                for row in range(current_row + 2, current_row + mat_rows + 3):
                    for col in range(1, mat_cols + 1):
                        cell = ws2.cell(row=row, column=col)
                        cell.border = thin_border
                
                current_row += mat_rows + 5
            
            # Add Box Costs
            if 'secondary_box_costs' in st.session_state and not st.session_state.secondary_box_costs.empty:
                ws2.cell(row=current_row, column=1, value="Table 2: Cardboard Box Cost").font = heading_font
                ws2.cell(row=current_row, column=1).fill = heading_fill
                
                st.session_state.secondary_box_costs.to_excel(writer, sheet_name='Secondary Calculations', index=False,
                                                       startrow=current_row + 1)
                
                # Apply borders
                box_rows_sec = len(st.session_state.secondary_box_costs)
                box_cols_sec = len(st.session_state.secondary_box_costs.columns)
                for row in range(current_row + 2, current_row + box_rows_sec + 3):
                    for col in range(1, box_cols_sec + 1):
                        cell = ws2.cell(row=row, column=col)
                        cell.border = thin_border
                
                current_row += box_rows_sec + 5
            
            # Add Polybag Costs
            if 'polybag_costs' in st.session_state and not st.session_state.polybag_costs.empty:
                ws2.cell(row=current_row, column=1, value="Table 3: Polybag Cost").font = heading_font
                ws2.cell(row=current_row, column=1).fill = heading_fill
                
                st.session_state.polybag_costs.to_excel(writer, sheet_name='Secondary Calculations', index=False,
                                                       startrow=current_row + 1)
                
                # Apply borders
                polybag_rows = len(st.session_state.polybag_costs)
                polybag_cols = len(st.session_state.polybag_costs.columns)
                for row in range(current_row + 2, current_row + polybag_rows + 3):
                    for col in range(1, polybag_cols + 1):
                        cell = ws2.cell(row=row, column=col)
                        cell.border = thin_border
                
                current_row += polybag_rows + 5
            
            # Add Stretchwrap Costs
            if 'stretchwrap_costs' in st.session_state and not st.session_state.stretchwrap_costs.empty:
                ws2.cell(row=current_row, column=1, value="Table 4: Stretch wrap cost").font = heading_font
                ws2.cell(row=current_row, column=1).fill = heading_fill
                
                st.session_state.stretchwrap_costs.to_excel(writer, sheet_name='Secondary Calculations', index=False,
                                                       startrow=current_row + 1)
                
                # Apply borders
                stretch_rows = len(st.session_state.stretchwrap_costs)
                stretch_cols = len(st.session_state.stretchwrap_costs.columns)
                for row in range(current_row + 2, current_row + stretch_rows + 3):
                    for col in range(1, stretch_cols + 1):
                        cell = ws2.cell(row=row, column=col)
                        cell.border = thin_border
                
                current_row += stretch_rows + 5
            
            # Add Crate/Pallet Cost Tables
            current_row += 2
            ws2.cell(row=current_row, column=1, value="Crate/Pallet Cost Tables").font = heading_font
            ws2.cell(row=current_row, column=1).fill = heading_fill
            current_row += 2
            
            # Add Crate Costs
            if 'crate_costs' in st.session_state and not st.session_state.crate_costs.empty:
                ws2.cell(row=current_row, column=1, value="Table 1: Crate Cost").font = heading_font
                ws2.cell(row=current_row, column=1).fill = heading_fill
                
                st.session_state.crate_costs.to_excel(writer, sheet_name='Secondary Calculations', index=False,
                                                     startrow=current_row + 1)
                
                # Apply borders
                crate_rows = len(st.session_state.crate_costs)
                crate_cols = len(st.session_state.crate_costs.columns)
                for row in range(current_row + 2, current_row + crate_rows + 3):
                    for col in range(1, crate_cols + 1):
                        cell = ws2.cell(row=row, column=col)
                        cell.border = thin_border
                
                current_row += crate_rows + 5
            
            # Add Pallet Costs
            if 'pallet_costs' in st.session_state and not st.session_state.pallet_costs.empty:
                ws2.cell(row=current_row, column=1, value="Table 2: Pallet Cost").font = heading_font
                ws2.cell(row=current_row, column=1).fill = heading_fill
                
                st.session_state.pallet_costs.to_excel(writer, sheet_name='Secondary Calculations', index=False,
                                                     startrow=current_row + 1)
                
                # Apply borders
                pallet_rows = len(st.session_state.pallet_costs)
                pallet_cols = len(st.session_state.pallet_costs.columns)
                for row in range(current_row + 2, current_row + pallet_rows + 3):
                    for col in range(1, pallet_cols + 1):
                        cell = ws2.cell(row=row, column=col)
                        cell.border = thin_border
                
                current_row += pallet_rows + 5
            
            # Add Strapping Clip Costs
            if 'strapping_clip_costs' in st.session_state and not st.session_state.strapping_clip_costs.empty:
                ws2.cell(row=current_row, column=1, value="Table 3: Strapping Clip Cost").font = heading_font
                ws2.cell(row=current_row, column=1).fill = heading_fill
                
                st.session_state.strapping_clip_costs.to_excel(writer, sheet_name='Secondary Calculations', index=False,
                                                             startrow=current_row + 1)
                
                # Apply borders
                clip_rows = len(st.session_state.strapping_clip_costs)
                clip_cols = len(st.session_state.strapping_clip_costs.columns)
                for row in range(current_row + 2, current_row + clip_rows + 3):
                    for col in range(1, clip_cols + 1):
                        cell = ws2.cell(row=row, column=col)
                        cell.border = thin_border
                
                current_row += clip_rows + 5
            
            # Add PP Strapping Costs
            if 'pp_strapping_costs' in st.session_state and not st.session_state.pp_strapping_costs.empty:
                ws2.cell(row=current_row, column=1, value="Table 4: PP Strapping Cost").font = heading_font
                ws2.cell(row=current_row, column=1).fill = heading_fill
                
                st.session_state.pp_strapping_costs.to_excel(writer, sheet_name='Secondary Calculations', index=False,
                                                           startrow=current_row + 1)
                
                # Apply borders
                pp_rows = len(st.session_state.pp_strapping_costs)
                pp_cols = len(st.session_state.pp_strapping_costs.columns)
                for row in range(current_row + 2, current_row + pp_rows + 3):
                    for col in range(1, pp_cols + 1):
                        cell = ws2.cell(row=row, column=col)
                        cell.border = thin_border
                
                current_row += pp_rows + 5
                
            # Add Special Comments Section for Secondary
            comments_start_row_sec = current_row + sec_calc_rows + 5
            ws2.cell(row=comments_start_row_sec, column=1, value="Special Comments under Secondary Packing").font = heading_font
            ws2.cell(row=comments_start_row_sec, column=1).fill = heading_fill
            
            # Get secondary comments from session state
            finish_secondary = st.session_state.get("finish_secondary", "Mill Finish")
            eco_friendly_secondary = st.session_state.get("eco_friendly_secondary", "Mac foam")
            
            # Determine protective tape comment
            if finish_secondary in ["PC", "WF", "Anodised"]:
                protective_tape_comment = "Protective tape required to avoid rejects"
            else:
                protective_tape_comment = "Protective tape is not mandatory"
            
            comments_text_sec = f"""**Packing Method Note:**
            
            1. Costing is done according to Secondary packing.
            
            2. The interleaving material is **"{eco_friendly_secondary}"**.
            
            3. {protective_tape_comment}
            
            4. Costing is inclusive of secondary packing - pallet or crate, however it is not inclusive of any labels or artwork. These will incur an additional charge."""
            
            # Add comments text
            current_row_sec = comments_start_row_sec + 2
            for line in comments_text_sec.split('\n'):
                ws2.cell(row=current_row_sec, column=1, value=line)
                current_row_sec += 1

    
            # Add Secondary Packing Cost Per Profile heading
            current_row += 2
            ws2.cell(row=current_row, column=1, value="Secondary Packing Cost Per Profile").font = heading_font
            ws2.cell(row=current_row, column=1).fill = heading_fill
            
            # Add Secondary Packing Cost Per Profile
            st.session_state.secondary_calculations.to_excel(writer, sheet_name='Secondary Calculations', index=False,
                                                  startrow=current_row + 2)
            
            # Apply borders to calculations table
            sec_calc_rows = len(st.session_state.secondary_calculations)
            sec_calc_cols = len(st.session_state.secondary_calculations.columns)
            for row in range(current_row + 3, current_row + sec_calc_rows + 4):
                for col in range(1, sec_calc_cols + 1):
                    cell = ws2.cell(row=row, column=col)
                    cell.border = thin_border
            
            # Add Crate/Pallet Dimensions if available
            current_row += sec_calc_rows + 6
            if 'crate_pallet_data' in st.session_state and not st.session_state.crate_pallet_data.empty:
                ws2.cell(row=current_row, column=1, value="Crate/Pallet Dimensions").font = heading_font
                ws2.cell(row=current_row, column=1).fill = heading_fill
                
                st.session_state.crate_pallet_data.to_excel(writer, sheet_name='Secondary Calculations', index=False,
                                                           startrow=current_row + 2)
                
                # Apply borders
                cp_dim_rows = len(st.session_state.crate_pallet_data)
                cp_dim_cols = len(st.session_state.crate_pallet_data.columns)
                for row in range(current_row + 3, current_row + cp_dim_rows + 4):
                    for col in range(1, cp_dim_cols + 1):
                        cell = ws2.cell(row=row, column=col)
                        cell.border = thin_border
                
                current_row += cp_dim_rows + 6
            
            # Add Crate/Pallet Cost Calculations if available
            if 'crate_pallet_calculations' in st.session_state and not st.session_state.crate_pallet_calculations.empty:
                ws2.cell(row=current_row, column=1, value="Crate/Pallet Cost Calculations").font = heading_font
                ws2.cell(row=current_row, column=1).fill = heading_fill
                
                st.session_state.crate_pallet_calculations.to_excel(writer, sheet_name='Secondary Calculations', index=False,
                                                     startrow=current_row + 2)
                
                # Apply borders
                cp_calc_rows = len(st.session_state.crate_pallet_calculations)
                cp_calc_cols = len(st.session_state.crate_pallet_calculations.columns)
                for row in range(current_row + 3, current_row + cp_calc_rows + 4):
                    for col in range(1, cp_calc_cols + 1):
                        cell = ws2.cell(row=row, column=col)
                        cell.border = thin_border

    
    
    # Get the Excel data
    output.seek(0)
    return output.getvalue()

# Add download button at the top
col1, col2, col3 = st.columns([1, 2, 1])
with col2:
    if st.button("📥 Download Complete Report (Excel)", type="primary", use_container_width=True):
        try:
            excel_data = create_excel_report()
            st.download_button(
                label="⬇️ Click to Download Excel File",
                data=excel_data,
                file_name="packing_costing_report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        except Exception as e:
            st.error(f"Error generating report: {str(e)}")

# Create tabs
tab1, tab2 = st.tabs(["Primary Calculations", "Secondary Calculations"])

with tab1:
    st.header("Primary Calculations")
    
    # Function to auto-calculate all fields in SKU table
    def auto_calculate_sku_table():
        """Auto-calculate all fields in SKU table"""
        if not st.session_state.primary_sku_data.empty:
            df_copy = st.session_state.primary_sku_data.copy()
            
            for idx, row in df_copy.iterrows():
                try:
                    # Calculate total weight per profile
                    unit_weight = float(row["Unit weight(kg/m)"])
                    length_mm = float(row["Length/mm"])
                    total_weight = unit_weight * (length_mm / 1000)
                    df_copy.at[idx, "total weight per profile (kg)"] = round(total_weight, 4)
                    
                    # Get profile dimensions
                    width = float(row["Width/mm"])
                    height = float(row["Height/mm"])
                    length = float(row["Length/mm"])
                    
                    # Calculate box dimensions (round UP to nearest 100)
                    def round_up_to_nearest_100(num):
                        return ((int(num) + 99) // 100) * 100
                    
                    # ALWAYS calculate box dimensions from profile dimensions (rounding up)
                    df_copy.at[idx, "Box Width/mm"] = round_up_to_nearest_100(width)
                    df_copy.at[idx, "Box Height/mm"] = round_up_to_nearest_100(height)
                    df_copy.at[idx, "Box Length/mm"] = round_up_to_nearest_100(length)
                    
                    # Get box dimensions
                    box_width = float(df_copy.at[idx, "Box Width/mm"])
                    box_height = float(df_copy.at[idx, "Box Height/mm"])
                    box_length = float(df_copy.at[idx, "Box Length/mm"])
                    
                    # Set default W/mm and H/mm if not set
                    if pd.isna(row["W/mm"]) or row["W/mm"] == "":
                        df_copy.at[idx, "W/mm"] = "Profiles are arranged in W direction"
                        df_copy.at[idx, "H/mm"] = "Profiles are arranged in height direction"
                    
                    # Calculate Number of profiles per box based on W/mm selection
                    w_direction = df_copy.at[idx, "W/mm"]
                    profiles_per_box = 0
                    
                    if w_direction == "Profiles are arranged in height direction":
                        # (Box Height/profile width)*(Box Width/profile Height)
                        if width > 0 and height > 0:
                            profiles_per_box = (box_height / width) * (box_width / height)
                    else:  # "Profiles are arranged in W direction"
                        # (Box Width/profile width)*(Box Height/profile Height)
                        if width > 0 and height > 0:
                            profiles_per_box = (box_width / width) * (box_height / height)
                    
                    # Use INTEGER value (floor) for number of profiles per box
                    df_copy.at[idx, "Number of profiles per box"] = int(profiles_per_box) if profiles_per_box > 0 else 0
                    
                    # Update H/mm based on W/mm selection
                    if w_direction == "Profiles are arranged in W direction":
                        df_copy.at[idx, "H/mm"] = "Profiles are arranged in height direction"
                    elif w_direction == "Profiles are arranged in height direction":
                        df_copy.at[idx, "H/mm"] = "Profiles are arranged in W direction"
                        
                except (ValueError, TypeError, ZeroDivisionError):
                    df_copy.at[idx, "Box Width/mm"] = 0
                    df_copy.at[idx, "Box Height/mm"] = 0
                    df_copy.at[idx, "Box Length/mm"] = 0
                    df_copy.at[idx, "Number of profiles per box"] = 0
            
            # Update session state
            st.session_state.primary_sku_data = df_copy
            st.success("Auto-calculation completed!")
            st.rerun()  # ADD THIS LINE to force immediate refresh
            
    
    # Function to auto-calculate box dimensions for ALL rows (including new ones)
    def auto_calculate_all_box_dimensions(df):
        """Auto-calculate box dimensions for all rows in dataframe"""
        df_copy = df.copy()
        for idx, row in df_copy.iterrows():
            try:
                # Get profile dimensions
                width = float(row["Width/mm"])
                height = float(row["Height/mm"])
                length = float(row["Length/mm"])
                
                # Calculate box dimensions (round UP to nearest 100)
                def round_up_to_nearest_100(num):
                    return ((int(num) + 99) // 100) * 100
                
                # ALWAYS calculate box dimensions from profile dimensions (rounding up)
                # This ensures ALL rows get calculated
                df_copy.at[idx, "Box Width/mm"] = round_up_to_nearest_100(width)
                df_copy.at[idx, "Box Height/mm"] = round_up_to_nearest_100(height)
                df_copy.at[idx, "Box Length/mm"] = round_up_to_nearest_100(length)
                
            except (ValueError, TypeError, ZeroDivisionError):
                # If there's an error, set default values
                df_copy.at[idx, "Box Width/mm"] = 0
                df_copy.at[idx, "Box Height/mm"] = 0
                df_copy.at[idx, "Box Length/mm"] = 0
        
        return df_copy
    
    # Sub topic 1 - SKU Table with dimensions with auto-calc button
    col1, col2 = st.columns([3, 1])
    with col1:
        st.subheader("SKU Table with dimensions")
    with col2:
        st.markdown("<br>", unsafe_allow_html=True)
        
        # Add a checkbox to control auto-calculation
        auto_calc_enabled = st.checkbox("Enable auto-calculation", value=True, key="auto_calc_enabled")
        
        if st.button("🔄 Auto-Calculate All Fields", 
                    help="Click to auto-calculate total weight, box dimensions, and profiles per box",
                    use_container_width=True):
            auto_calculate_sku_table()
            st.rerun()
    
    # Initialize session state for primary SKU data with new columns
    if 'primary_sku_data' not in st.session_state:
        st.session_state.primary_sku_data = pd.DataFrame(columns=[
            "SKU No", 
            "Unit weight(kg/m)", 
            "total weight per profile (kg)", 
            "Width/mm", 
            "Height/mm", 
            "Length/mm",
            "Box Width/mm",
            "Box Height/mm", 
            "Box Length/mm",
            "W/mm",
            "H/mm",
            "Number of profiles per box",
            "Comment on fabrication"
        ])
    
    # Function to calculate box dimensions and number of profiles per box
    # Function to calculate box dimensions and number of profiles per box
    def calculate_box_and_profiles(df):
        """Calculate box dimensions and number of profiles per box"""
        df_copy = df.copy()
        for idx, row in df_copy.iterrows():
            try:
                # Get profile dimensions
                width = float(row["Width/mm"])
                height = float(row["Height/mm"])
                length = float(row["Length/mm"])
                
                # Calculate box dimensions (round UP to nearest 100)
                def round_up_to_nearest_100(num):
                    """Round UP to nearest 100"""
                    return ((int(num) + 99) // 100) * 100
                
                # ALWAYS calculate box dimensions from profile dimensions (rounding up)
                # But only if box dimensions are 0 or not set, OR if they match the profile dimensions
                # (this allows users to manually edit them)
                current_box_width = float(row["Box Width/mm"]) if not pd.isna(row["Box Width/mm"]) else 0
                current_box_height = float(row["Box Height/mm"]) if not pd.isna(row["Box Height/mm"]) else 0
                current_box_length = float(row["Box Length/mm"]) if not pd.isna(row["Box Length/mm"]) else 0
                
                # Auto-calculate only if box dimensions are not already manually set
                # (assume manually set if they're not exactly the rounded up value of profile dimensions)
                expected_box_width = round_up_to_nearest_100(width)
                expected_box_height = round_up_to_nearest_100(height)
                expected_box_length = round_up_to_nearest_100(length)
                
                if current_box_width == 0 or current_box_width == expected_box_width:
                    df_copy.at[idx, "Box Width/mm"] = expected_box_width
                else:
                    # Keep user's manual entry
                    df_copy.at[idx, "Box Width/mm"] = current_box_width
                    
                if current_box_height == 0 or current_box_height == expected_box_height:
                    df_copy.at[idx, "Box Height/mm"] = expected_box_height
                else:
                    # Keep user's manual entry
                    df_copy.at[idx, "Box Height/mm"] = current_box_height
                    
                if current_box_length == 0 or current_box_length == expected_box_length:
                    df_copy.at[idx, "Box Length/mm"] = expected_box_length
                else:
                    # Keep user's manual entry
                    df_copy.at[idx, "Box Length/mm"] = current_box_length
                
                # Get box dimensions (either auto-calculated or manually entered)
                box_width = float(df_copy.at[idx, "Box Width/mm"])
                box_height = float(df_copy.at[idx, "Box Height/mm"])
                box_length = float(df_copy.at[idx, "Box Length/mm"])
                
                # Set default W/mm and H/mm if not set
                if pd.isna(row["W/mm"]) or row["W/mm"] == "":
                    df_copy.at[idx, "W/mm"] = "Profiles are arranged in W direction"
                    df_copy.at[idx, "H/mm"] = "Profiles are arranged in height direction"
                
                # Calculate Number of profiles per box based on W/mm selection
                w_direction = df_copy.at[idx, "W/mm"]
                profiles_per_box = 0
                
                if w_direction == "Profiles are arranged in height direction":
                    # (Box Height/profile width)*(Box Width/profile Height)
                    if width > 0 and height > 0:
                        profiles_per_box = (box_height / width) * (box_width / height)
                else:  # "Profiles are arranged in W direction"
                    # (Box Width/profile width)*(Box Height/profile Height)
                    if width > 0 and height > 0:
                        profiles_per_box = (box_width / width) * (box_height / height)
                
                # Use INTEGER value (floor) for number of profiles per box
                df_copy.at[idx, "Number of profiles per box"] = int(profiles_per_box) if profiles_per_box > 0 else 0
                
                # Update H/mm based on W/mm selection
                if w_direction == "Profiles are arranged in W direction":
                    df_copy.at[idx, "H/mm"] = "Profiles are arranged in height direction"
                elif w_direction == "Profiles are arranged in height direction":
                    df_copy.at[idx, "H/mm"] = "Profiles are arranged in W direction"
                    
            except (ValueError, TypeError, ZeroDivisionError):
                df_copy.at[idx, "Box Width/mm"] = 0
                df_copy.at[idx, "Box Height/mm"] = 0
                df_copy.at[idx, "Box Length/mm"] = 0
                df_copy.at[idx, "Number of profiles per box"] = 0
        
        return df_copy
    
    # Calculate initial values
    if not st.session_state.primary_sku_data.empty:
        st.session_state.primary_sku_data = calculate_box_and_profiles(st.session_state.primary_sku_data)
    
    # Create editable dataframe for SKU input
    edited_sku_df = st.data_editor(
        st.session_state.primary_sku_data,
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "SKU No": st.column_config.TextColumn("SKU No", required=True),
            "Unit weight(kg/m)": st.column_config.NumberColumn("Unit weight(kg/m)", required=True, min_value=0, format="%.4f"),
            "total weight per profile (kg)": st.column_config.NumberColumn("total weight per profile (kg)", required=True, min_value=0, format="%.4f", disabled=True),
            "Width/mm": st.column_config.NumberColumn("Width/mm", required=True, min_value=0, format="%.1f"),
            "Height/mm": st.column_config.NumberColumn("Height/mm", required=True, min_value=0, format="%.1f"),
            "Length/mm": st.column_config.NumberColumn("Length/mm", required=True, min_value=0, format="%.1f"),
            "Box Width/mm": st.column_config.NumberColumn("Box Width/mm", required=True, min_value=0, format="%d", step=100),
            "Box Height/mm": st.column_config.NumberColumn("Box Height/mm", required=True, min_value=0, format="%d", step=100),
            "Box Length/mm": st.column_config.NumberColumn("Box Length/mm", required=True, min_value=0, format="%d", step=100),
            "W/mm": st.column_config.SelectboxColumn(
                "W/mm",
                options=[
                    "Profiles are arranged in W direction", 
                    "Profiles are arranged in height direction"
                ],
                required=True
            ),
            "H/mm": st.column_config.TextColumn(
                "H/mm",
                required=True,
                disabled=True
            ),
            "Number of profiles per box": st.column_config.NumberColumn(
                "Number of profiles per box", 
                required=True, 
                min_value=0, 
                format="%d",  # Changed from "%.2f" to "%d" for integer display
                disabled=True
            ),
            "Comment on fabrication": st.column_config.SelectboxColumn(
                "Comment on fabrication",
                options=["Fabricated", "Just Cutting"],
                required=True
            )
        },
        key="sku_editor_primary"
    )
    
    # Callback function to auto-calculate total weight
    def calculate_total_weight(df):
        """Calculate total weight per profile based on unit weight and length"""
        df_copy = df.copy()
        for idx, row in df_copy.iterrows():
            try:
                unit_weight = float(row["Unit weight(kg/m)"])
                length_mm = float(row["Length/mm"])
                # Calculate: Unit weight(kg/m) * (Length(mm) / 1000)
                total_weight = unit_weight * (length_mm / 1000)
                df_copy.at[idx, "total weight per profile (kg)"] = round(total_weight, 4)
            except (ValueError, TypeError):
                df_copy.at[idx, "total weight per profile (kg)"] = 0
        return df_copy
    
    # Update the session state with calculated weights and box dimensions
    if not edited_sku_df.equals(st.session_state.primary_sku_data):
        # Only auto-calculate if enabled
        if st.session_state.get("auto_calc_enabled", True):
            # First, ensure ALL box dimensions are calculated (including new rows)
            edited_sku_df = auto_calculate_all_box_dimensions(edited_sku_df)
            
            # Recalculate total weights
            edited_sku_df_with_weights = calculate_total_weight(edited_sku_df)
            # Recalculate box dimensions and profiles (this includes profiles per box calculation)
            edited_sku_df_with_all = calculate_box_and_profiles(edited_sku_df_with_weights)
            st.session_state.primary_sku_data = edited_sku_df_with_all
            st.rerun()
    
    st.divider()
    
    # Section 2: Common Packing selections
    st.subheader("Common Packing Selections")
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        finish = st.selectbox(
            "Finish",
            ["Mill Finish", "Anodised", "PC", "WF"],
            key="finish_primary"
        )
    
    with col2:
        interleaving_required = st.selectbox(
            "Interleaving Required",
            ["Yes", "No"],
            key="interleaving_primary"
        )
    
    with col3:
        # Get current finish selection
        finish = st.session_state.get("finish_primary", "Mill Finish")
        
        # Define available options based on finish
        if finish == "Anodised":
            eco_options = ["Mac foam", "Stretch wrap"]  # No Craft Paper for Anodised
        else:
            eco_options = ["Mac foam", "Stretch wrap", "Craft Paper"]
        
        eco_friendly = st.selectbox(
            "Eco-Friendly Packing Material",
            options=eco_options,
            key="eco_friendly_primary"
        )
    
    with col4:
        protective_tape = st.selectbox(
            "Protective Tape (Customer Specified)",
            ["Yes", "No"],
            key="protective_tape_primary"
        )
    
    st.divider()
    
    # Section 3: Primary Packing Total Cost
    st.subheader("Primary Packing Total Cost")
    
    # Table 1: Material Costs
    st.markdown("**Table 1: Primary Packing Material Costs**")
    
    # Initialize session state for material costs
    if 'primary_material_costs' not in st.session_state:
        st.session_state.primary_material_costs = pd.DataFrame({
            "Material": ["McFoam", "Craft Paper", "Protective Tape", "Stretchwrap"],
            "Cost/ m²": [51.00, 34.65, 100.65, 14.38]
        })
    
    # Create editable material costs table
    edited_material_df = st.data_editor(
        st.session_state.primary_material_costs,
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "Material": st.column_config.TextColumn("Material", required=True),
            "Cost/ m²": st.column_config.NumberColumn("Cost/ m²", required=True, min_value=0, format="%.2f")
        },
        key="material_editor_primary"
    )
    
    # Update session state
    st.session_state.primary_material_costs = edited_material_df
    
    # Table 2: Cardboard Box Cost
    if box_ply == "2 ply":
        # 2 ply box costs (you can adjust these values)
        default_box_costs = pd.DataFrame({
            "Length(mm)": [330],
            "Width (mm)": [210],
            "Height (mm)": [135],
            "Cost (LKR)": [150.00]  # Lower cost for 2 ply
        })
    else:  # 3 ply
        default_box_costs = pd.DataFrame({
            "Length(mm)": [330],
            "Width (mm)": [210],
            "Height (mm)": [135],
            "Cost (LKR)": [205.00]  # Higher cost for 3 ply
        })
    
    # Update session state if ply selection changed
    if 'previous_box_ply' not in st.session_state or st.session_state.previous_box_ply != box_ply:
        st.session_state.primary_box_costs = default_box_costs.copy()
        st.session_state.previous_box_ply = box_ply
    
    edited_box_df = st.data_editor(
        st.session_state.primary_box_costs,
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "Length(mm)": st.column_config.NumberColumn("Length(mm)", required=True, min_value=0),
            "Width (mm)": st.column_config.NumberColumn("Width (mm)", required=True, min_value=0),
            "Height (mm)": st.column_config.NumberColumn("Height (mm)", required=True, min_value=0),
            "Cost (LKR)": st.column_config.NumberColumn("Cost (LKR)", required=True, min_value=0, format="%.2f")
        },
        key="box_editor_primary"
    )
    
    # Update session state
    st.session_state.primary_box_costs = edited_box_df
    
    # Add Calculate button
    st.markdown("---")
    calculate_col1, calculate_col2, calculate_col3 = st.columns([1, 1, 1])
    with calculate_col2:
        calculate_primary_btn = st.button("🔢 Calculate Primary Packing Costs", 
                                         type="primary", 
                                         use_container_width=True,
                                         on_click=trigger_primary_calculation)
    
    # Calculate and display Table 3: Primary Packing Total Cost
    if st.session_state.calculate_primary:
        st.markdown("**Table 3: Primary Packing Total Cost**")
        
        if not st.session_state.primary_sku_data.empty:
            # Prepare calculations
            calculations_data = []
            
            # Get material costs as dictionary
            material_cost_dict = dict(zip(
                st.session_state.primary_material_costs["Material"],
                st.session_state.primary_material_costs["Cost/ m²"]
            ))
            
            # Get reference box dimensions and cost
            ref_box = st.session_state.primary_box_costs.iloc[0]
            
            for _, sku in st.session_state.primary_sku_data.iterrows():
                try:
                    # Get SKU dimensions
                    width = float(sku["Width/mm"])
                    height = float(sku["Height/mm"])
                    length = float(sku["Length/mm"])
                    unit_weight = float(sku["Unit weight(kg/m)"])
                    total_weight = float(sku["total weight per profile (kg)"])
                    
                    # Get box dimensions from SKU table
                    box_width = float(sku["Box Width/mm"])
                    box_height = float(sku["Box Height/mm"])
                    box_length = float(sku["Box Length/mm"])
                    profiles_per_box = float(sku["Number of profiles per box"])
                    
                    # Calculate Surface Area in m² (using profile dimensions)
                    sa_m2 = (2 * ((width * height) + (width * length) + (height * length))) / (1000 * 1000)
                    
                    # Calculate Interleaving Cost
                    interleaving_cost = 0
                    if interleaving_required == "Yes":
                        material_map = {
                            "Mac foam": "McFoam",
                            "Stretch wrap": "Stretchwrap",
                            "Craft Paper": "Craft Paper"
                        }
                        selected_material = material_map.get(eco_friendly, "McFoam")
                        cost_per_m2 = material_cost_dict.get(selected_material, 0)
                        interleaving_cost = (cost_per_m2 * sa_m2)/profiles_per_box
                    
                    # Calculate Protective Tape Cost
                    protective_tape_cost = 0
                    if protective_tape == "Yes":
                        cost_per_m2 = material_cost_dict.get("Protective Tape", 0)
                        protective_tape_cost = (cost_per_m2 * sa_m2)/profiles_per_box
                    
                    # Calculate Packing Cost - UPDATED FORMULA
                    # Using box dimensions from SKU table divided by number of profiles per box
                    packing_cost = 0
                    # In the primary calculations section where packing_cost is calculated:
                    if ref_box["Length(mm)"] > 0 and ref_box["Width (mm)"] > 0 and ref_box["Height (mm)"] > 0:
                        if profiles_per_box > 0:
                            # Calculate cost per box volume
                            box_volume_cost = (ref_box["Cost (LKR)"] / (ref_box["Width (mm)"] * ref_box["Height (mm)"] * ref_box["Length(mm)"]))
                            # Calculate actual box volume from SKU table
                            actual_box_volume = box_width * box_height * box_length
                            # Total box cost = cost per volume * actual box volume
                            total_box_cost = box_volume_cost * actual_box_volume
                            # Packing cost per profile = total box cost / number of profiles per box
                            packing_cost = total_box_cost / profiles_per_box  # This now uses integer value
                        else:
                            # Fallback to original calculation if profiles_per_box is 0
                            packing_cost = (ref_box["Cost (LKR)"] / (ref_box["Width (mm)"] * ref_box["Height (mm)"] * ref_box["Length(mm)"])) * (width * height * length)
                            
                    # Calculate Total Cost
                    total_cost = interleaving_cost + protective_tape_cost + packing_cost

                    # Calculate Cost/kg (LKR)
                    cost_per_kg = total_cost / total_weight if total_weight > 0 else 0

                    calculations_data.append({
                        "SKU": sku["SKU No"],
                        "Interleaving cost": round(interleaving_cost, 2),
                        "Protective tape cost": round(protective_tape_cost, 2),
                        "Packing type": "Cardboard box",
                        "Profiles per box": int(profiles_per_box),
                        "Packing Cost (LKR)": round(packing_cost, 2),
                        "Total Cost per profile/LKR": round(total_cost, 2),
                        "Cost/kg (LKR)": round(cost_per_kg, 2)
                    })
                    
                except (ValueError, TypeError, ZeroDivisionError):
                    continue
            
            if calculations_data:
                # Store calculations in session state
                st.session_state.primary_calculations = pd.DataFrame(calculations_data)
                
                # Create editable dataframe with initial calculations
                editable_df = st.session_state.primary_calculations.copy()
                
                # Create editable dataframe for calculations
                edited_calc_df = st.data_editor(
                    editable_df,
                    num_rows="fixed",
                    use_container_width=True,
                    column_config={
                        "SKU": st.column_config.TextColumn("SKU", required=True, disabled=True),
                        "Interleaving cost": st.column_config.NumberColumn("Interleaving cost", required=True, min_value=0, format="%.2f", disabled=True),
                        "Protective tape cost": st.column_config.NumberColumn("Protective tape cost", required=True, min_value=0, format="%.2f", disabled=True),
                        "Packing type": st.column_config.TextColumn("Packing type", required=True, disabled=True),
                        "Profiles per box": st.column_config.NumberColumn("Profiles per box", required=True, min_value=0, format="%.2f", disabled=True),
                        "Packing Cost (LKR)": st.column_config.NumberColumn("Packing Cost (LKR)", required=True, min_value=0, format="%.2f", disabled=True),
                        "Total Cost per profile/LKR": st.column_config.NumberColumn("Total Cost per profile/LKR", required=True, min_value=0, format="%.2f", disabled=True),
                        "Cost/kg (LKR)": st.column_config.NumberColumn("Cost/kg (LKR)", required=True, min_value=0, format="%.2f", disabled=True)
                    },
                    key="primary_calculations_editor"
                )
                
                # Recalculate if box dimensions changed
                if not edited_calc_df.equals(st.session_state.primary_calculations):
                    # Recalculate packing costs based on new dimensions
                    ref_length = ref_box["Length(mm)"]
                    ref_width = ref_box["Width (mm)"]
                    ref_height = ref_box["Height (mm)"]
                    ref_cost = ref_box["Cost (LKR)"]
                    
                    if ref_length > 0 and ref_width > 0 and ref_height > 0:
                        cost_per_volume = ref_cost / (ref_width * ref_height * ref_length)
                    else:
                        cost_per_volume = 0
                    
                    # Update calculations
                    for idx, row in edited_calc_df.iterrows():
                        try:
                            # Get profiles per box from original SKU data
                            sku_no = row["SKU"]
                            sku_row = st.session_state.primary_sku_data[st.session_state.primary_sku_data["SKU No"] == sku_no]
                            if not sku_row.empty:
                                profiles_per_box = float(sku_row.iloc[0]["Number of profiles per box"])
                            else:
                                profiles_per_box = 1
                            
                            box_width = float(row["Box width/mm"])
                            box_height = float(row["Box height/mm"])
                            box_length = float(row["Box length/mm"])
                            
                            box_volume = box_width * box_height * box_length
                            
                            if cost_per_volume > 0 and profiles_per_box > 0:
                                new_packing_cost = (cost_per_volume * box_volume) / profiles_per_box
                            else:
                                new_packing_cost = 0
                            
                            # Get original costs
                            interleaving_cost = float(row["Interleaving cost"])
                            protective_tape_cost = float(row["Protective tape cost"])
                            
                            # Calculate new total cost
                            new_total_cost = interleaving_cost + protective_tape_cost + new_packing_cost
                            
                            # Update the row
                            edited_calc_df.at[idx, "Packing Cost (LKR)"] = round(new_packing_cost, 2)
                            edited_calc_df.at[idx, "Total Cost"] = round(new_total_cost, 2)
                            
                        except (ValueError, TypeError):
                            continue
                    
                    # Update session state
                    st.session_state.primary_calculations = edited_calc_df
                    st.rerun()
                
                # Display the updated dataframe
                st.dataframe(st.session_state.primary_calculations, use_container_width=True)
                
            else:
                st.warning("Enter valid SKU data")
        else:
            st.info("Enter SKU data")
        
        # Reset calculation flag after displaying
        st.session_state.calculate_primary = False
    
    st.divider()
    
    # Section 4: Special Comments Section
    st.subheader("Special Comments Section")
    
    comments_box = f"""
    Costing is done according to primary packing. Therefore, this cost does not include any crate or palletizing charges. Please note that secondary packaging will incur an additional charge.

    The interleaving material is "{eco_friendly}".

    Protective tape required to avoid rejects

    Costing is only inclusive of interleaving required & Cardboard Box/Polybag.
    """
    
    st.info(comments_box)

# Note: Secondary Calculations tab code remains unchanged from the original
with tab2:
    st.header("Secondary Calculations")
    
    # SKU input table
    st.subheader("SKU Table with dimensions")
    
    # Initialize session state for secondary SKU data
    if 'secondary_sku_data' not in st.session_state:
        st.session_state.secondary_sku_data = pd.DataFrame(columns=["SKU No", "Width/mm", "Height/mm", "Length/mm", "Comment on fabrication"])
    
    edited_sku_df_tab2 = st.data_editor(
        st.session_state.secondary_sku_data,
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "SKU No": st.column_config.TextColumn("SKU No", required=True),
            "Width/mm": st.column_config.NumberColumn("Width/mm", required=True, min_value=0, format="%.1f"),  # Changed to allow decimals
            "Height/mm": st.column_config.NumberColumn("Height/mm", required=True, min_value=0, format="%.1f"),  # Changed to allow decimals
            "Length/mm": st.column_config.NumberColumn("Length/mm", required=True, min_value=0, format="%.1f"),  # Changed to allow decimals
            "Comment on fabrication": st.column_config.SelectboxColumn(
                "Comment on fabrication",
                options=["Fabricated", "Just Cutting"],
                required=True
            )
        },
        key="sku_editor_secondary"
    )
    
    # Update session state
    st.session_state.secondary_sku_data = edited_sku_df_tab2
    
    st.divider()
    
    # Common packing selections
    st.subheader("Common Packing Selections")
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        finish_tab2 = st.selectbox(
            "Finish",
            ["Mill Finish", "Anodised", "PC", "WF"],
            key="finish_secondary"
        )
    
    with col2:
        interleaving_required_tab2 = st.selectbox(
            "Interleaving Required",
            ["Yes", "No"],
            key="interleaving_secondary"
        )
    
    with col3:
        eco_friendly_tab2 = st.selectbox(
            "Eco-Friendly Packing Material",
            ["Mac foam", "Stretch wrap", "Craft Paper"],
            key="eco_friendly_secondary"
        )
    
    with col4:
        protective_tape_tab2 = st.selectbox(
            "Protective Tape (Customer Specified)",
            ["Yes", "No"],
            key="protective_tape_secondary"
        )
    
    st.divider()
    
    # New Section: Input Bundling Data
    st.subheader("Input Bundling Data")
    
    # Section 1 - Number of layers (Method 1)
    st.markdown("**Section 1 - Number of layers (Method 1)**")
    
    # Initialize bundling data in session state
    if 'bundling_data' not in st.session_state:
        st.session_state.bundling_data = pd.DataFrame(columns=[
            "Number of rows/bundle", 
            "Number of layer/bundle", 
            "width prof.type", 
            "height prof.type"
        ])
    
    # Create editable bundling data table
    edited_bundling_df = st.data_editor(
        st.session_state.bundling_data,
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "Number of rows/bundle": st.column_config.NumberColumn("Number of rows/bundle", required=True, min_value=1),
            "Number of layer/bundle": st.column_config.NumberColumn("Number of layer/bundle", required=True, min_value=1),
            "width prof.type": st.column_config.SelectboxColumn(
                "width prof.type",
                options=["W/mm", "H/mm"],
                required=True
            ),
            "height prof.type": st.column_config.TextColumn("height prof.type", required=True, disabled=True)
        },
        key="bundling_editor"
    )
    
    # Update height prof.type based on width prof.type selection
    updated_bundling_df = edited_bundling_df.copy()
    if not updated_bundling_df.empty:
        for idx in range(len(updated_bundling_df)):
            width_prof = updated_bundling_df.iloc[idx]["width prof.type"]
            if width_prof == "W/mm":
                updated_bundling_df.at[idx, "height prof.type"] = "H/mm"
            elif width_prof == "H/mm":
                updated_bundling_df.at[idx, "height prof.type"] = "W/mm"
    
    st.session_state.bundling_data = updated_bundling_df
    
    st.divider()
    
    # Section 2 - Size of bundle (Method 2)
    st.markdown("**Section 2 - Size of bundle (Method 2)**")
    
    # Initialize bundle size data in session state
    if 'bundle_size_data' not in st.session_state:
        st.session_state.bundle_size_data = pd.DataFrame(columns=[
            "Bundle width/mm", 
            "Bundle Height/mm"
        ])
    
    # Create editable bundle size table
    edited_bundle_size_df = st.data_editor(
        st.session_state.bundle_size_data,
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "Bundle width/mm": st.column_config.NumberColumn("Bundle width/mm", required=True, min_value=0, format="%.1f"),  # Changed to allow decimals
            "Bundle Height/mm": st.column_config.NumberColumn("Bundle Height/mm", required=True, min_value=0, format="%.1f")  # Changed to allow decimals
        },
        key="bundle_size_editor"
    )
    
    # Update session state
    st.session_state.bundle_size_data = edited_bundle_size_df
    
    st.divider()
    
    # New Section: Secondary Packing Cost (Per Profile)
    st.subheader("Secondary Packing Cost (Per Profile)")
    
    # Table 1: Primary Packing Material Costs
    st.markdown("**Table 1: Primary Packing Material Costs**")
    
    # Initialize secondary material costs in session state
    if 'secondary_material_costs' not in st.session_state:
        st.session_state.secondary_material_costs = pd.DataFrame({
            "Material": ["McFoam", "Craft Paper", "Protective Tape", "Stretchwrap"],
            "Cost/ m²": [51.00, 34.65, 100.65, 14.38]
        })
    
    # Create editable material costs table
    edited_secondary_material_df = st.data_editor(
        st.session_state.secondary_material_costs,
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "Material": st.column_config.TextColumn("Material", required=True),
            "Cost/ m²": st.column_config.NumberColumn("Cost/ m²", required=True, min_value=0, format="%.2f")
        },
        key="secondary_material_editor"
    )
    
    # Update session state
    st.session_state.secondary_material_costs = edited_secondary_material_df
    
    # Table 2: Cardboard Box Cost
# Table 2: Cardboard Box Cost
st.markdown("**Table 2: Cardboard Box Cost**")
col1, col2 = st.columns([1, 2])
with col1:
    box_ply = st.selectbox(
        "Select Box Ply",
        ["2 ply", "3 ply"],
        key="box_ply"
    )

# Initialize session state for box costs based on ply selection
if 'primary_box_costs' not in st.session_state:
    # Default to 3 ply costs
    st.session_state.primary_box_costs = pd.DataFrame({
        "Length(mm)": [330],
        "Width (mm)": [210],
        "Height (mm)": [135],
        "Cost (LKR)": [205.00]
    })

# Define box costs based on ply selection
if box_ply == "2 ply":
    # 2 ply box costs (you can adjust these values)
    default_box_costs = pd.DataFrame({
        "Length(mm)": [330],
        "Width (mm)": [210],
        "Height (mm)": [135],
        "Cost (LKR)": [150.00]  # Lower cost for 2 ply
    })
else:  # 3 ply
    default_box_costs = pd.DataFrame({
        "Length(mm)": [330],
        "Width (mm)": [210],
        "Height (mm)": [135],
        "Cost (LKR)": [205.00]  # Higher cost for 3 ply
    })

# Update session state if ply selection changed
if 'previous_box_ply' not in st.session_state or st.session_state.get("previous_box_ply") != box_ply:
    st.session_state.primary_box_costs = default_box_costs.copy()
    st.session_state.previous_box_ply = box_ply

# Create editable box costs table
edited_box_df = st.data_editor(
    st.session_state.primary_box_costs,
    num_rows="dynamic",
    use_container_width=True,
    column_config={
        "Length(mm)": st.column_config.NumberColumn("Length(mm)", required=True, min_value=0),
        "Width (mm)": st.column_config.NumberColumn("Width (mm)", required=True, min_value=0),
        "Height (mm)": st.column_config.NumberColumn("Height (mm)", required=True, min_value=0),
        "Cost (LKR)": st.column_config.NumberColumn("Cost (LKR)", required=True, min_value=0, format="%.2f")
    },
    key="box_editor_primary"
)

# Update session state
st.session_state.primary_box_costs = edited_box_df
    
    # Table 3: Polybag Cost
st.markdown("**Table 3: Polybag Cost**")
    
    # Initialize polybag costs in session state
    if 'polybag_costs' not in st.session_state:
        st.session_state.polybag_costs = pd.DataFrame({
            "Polybag size (inches)": [9],
            "Cost/m (LKR/m)": [12.8]
        })
    
    # Create editable polybag costs table
    edited_polybag_df = st.data_editor(
        st.session_state.polybag_costs,
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "Polybag size (inches)": st.column_config.NumberColumn("Polybag size (inches)", required=True, min_value=0),
            "Cost/m (LKR/m)": st.column_config.NumberColumn("Cost/m (LKR/m)", required=True, min_value=0, format="%.2f")
        },
        key="polybag_editor"
    )
    
    # Update session state
    st.session_state.polybag_costs = edited_polybag_df
    
    # Table 4: Stretch wrap cost
    st.markdown("**Table 4: Stretch wrap cost**")
    
    # Initialize stretch wrap costs in session state
    if 'stretchwrap_costs' not in st.session_state:
        st.session_state.stretchwrap_costs = pd.DataFrame({
            "Area (mm²)": [210000],
            "Cost (LKR/mm²)": [135]
        })
    
    # Create editable stretch wrap costs table
    edited_stretchwrap_df = st.data_editor(
        st.session_state.stretchwrap_costs,
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "Area (mm²)": st.column_config.NumberColumn("Area (mm²)", required=True, min_value=0),
            "Cost (LKR/mm²)": st.column_config.NumberColumn("Cost (LKR/mm²)", required=True, min_value=0, format="%.10f")
        },
        key="stretchwrap_editor"
    )
    
    # Update session state
    st.session_state.stretchwrap_costs = edited_stretchwrap_df
    
    # Add packing type selection
    packing_type = st.selectbox(
        "Select Packing Type",
        ["polybag", "cardboard box"],
        key="secondary_packing_type"
    )
    
    # New Section: Crate/Pallet dimensions table
    st.subheader("Crate/Pallet Dimensions Table")
    
    # Initialize crate/pallet data in session state
    if 'crate_pallet_data' not in st.session_state:
        st.session_state.crate_pallet_data = pd.DataFrame(columns=[
            "SKU", 
            "packing method", 
            "Width/mm", 
            "Height/mm", 
            "Length/mm"
        ])
    
    # Create dataframe with SKUs from the SKU input table
    if not st.session_state.secondary_sku_data.empty:
        # Get unique SKUs from the SKU table
        sku_list = st.session_state.secondary_sku_data["SKU No"].unique().tolist()
        
        # Check if crate_pallet_data needs to be updated with new SKUs
        existing_skus = st.session_state.crate_pallet_data["SKU"].unique().tolist() if not st.session_state.crate_pallet_data.empty else []
        
        # Add new SKUs that aren't already in the crate/pallet table
        new_skus = [sku for sku in sku_list if sku not in existing_skus]
        
        if new_skus:
            new_rows = []
            for sku in new_skus:
                new_rows.append({
                    "SKU": sku,
                    "packing method": "pallet",  # default value
                    "Width/mm": 0,
                    "Height/mm": 0,
                    "Length/mm": 0
                })
            
            if new_rows:
                new_df = pd.DataFrame(new_rows)
                st.session_state.crate_pallet_data = pd.concat([st.session_state.crate_pallet_data, new_df], ignore_index=True)
    
    # Create editable crate/pallet dimensions table
    edited_crate_pallet_df = st.data_editor(
        st.session_state.crate_pallet_data,
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "SKU": st.column_config.TextColumn("SKU", required=True),
            "packing method": st.column_config.SelectboxColumn(
                "packing method",
                options=["pallet", "crate"],
                required=True
            ),
            "Width/mm": st.column_config.NumberColumn("Width/mm", required=True, min_value=0, format="%.1f"),  # Changed to allow decimals
            "Height/mm": st.column_config.NumberColumn("Height/mm", required=True, min_value=0, format="%.1f"),  # Changed to allow decimals
            "Length/mm": st.column_config.NumberColumn("Length/mm", required=True, min_value=0, format="%.1f")  # Changed to allow decimals
        },
        key="crate_pallet_editor"
    )
    
    # Update session state
    st.session_state.crate_pallet_data = edited_crate_pallet_df
    
    # New Section: Total crate/pallet cost
    st.subheader("Total Crate/Pallet Cost")
    
    # Table 1: Crate cost
    st.markdown("**Table 1: Crate Cost**")
    
    # Initialize crate costs in session state
    if 'crate_costs' not in st.session_state:
        st.session_state.crate_costs = pd.DataFrame({
            "Crate width/mm": [480],
            "Crate Height/mm": [590],
            "Crate Length/mm": [2000],
            "Cost (LKR)": [5000.00]
        })
    
    # Create editable crate costs table
    edited_crate_costs_df = st.data_editor(
        st.session_state.crate_costs,
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "Crate width/mm": st.column_config.NumberColumn("Crate width/mm", required=True, min_value=0),
            "Crate Height/mm": st.column_config.NumberColumn("Crate Height/mm", required=True, min_value=0),
            "Crate Length/mm": st.column_config.NumberColumn("Crate Length/mm", required=True, min_value=0),
            "Cost (LKR)": st.column_config.NumberColumn("Cost (LKR)", required=True, min_value=0, format="%.2f")
        },
        key="crate_costs_editor"
    )
    
    # Update session state
    st.session_state.crate_costs = edited_crate_costs_df
    
    # Table 2: Pallet Cost
    st.markdown("**Table 2: Pallet Cost**")
    
    # Initialize pallet costs in session state
    if 'pallet_costs' not in st.session_state:
        st.session_state.pallet_costs = pd.DataFrame({
            "Pallet width/mm": [2000],
            "Pallet Height/mm": [600],
            "Cost (LKR)": [3000.00]
        })
    
    # Create editable pallet costs table
    edited_pallet_costs_df = st.data_editor(
        st.session_state.pallet_costs,
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "Pallet width/mm": st.column_config.NumberColumn("Pallet width/mm", required=True, min_value=0),
            "Pallet Height/mm": st.column_config.NumberColumn("Pallet Height/mm", required=True, min_value=0),
            "Cost (LKR)": st.column_config.NumberColumn("Cost (LKR)", required=True, min_value=0, format="%.2f")
        },
        key="pallet_costs_editor"
    )
    
    # Update session state
    st.session_state.pallet_costs = edited_pallet_costs_df
    
    # Table 3: Strapping clip cost
    st.markdown("**Table 3: Strapping Clip Cost**")
    
    # Initialize strapping clip costs in session state
    if 'strapping_clip_costs' not in st.session_state:
        st.session_state.strapping_clip_costs = pd.DataFrame({
            "number of clips": [1],
            "Cost": [12.00]
        })
    
    # Create editable strapping clip costs table
    edited_strapping_clip_df = st.data_editor(
        st.session_state.strapping_clip_costs,
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "number of clips": st.column_config.NumberColumn("number of clips", required=True, min_value=1),
            "Cost": st.column_config.NumberColumn("Cost", required=True, min_value=0, format="%.2f")
        },
        key="strapping_clip_editor"
    )
    
    # Update session state
    st.session_state.strapping_clip_costs = edited_strapping_clip_df
    
    # Table 4: PP strapping cost
    st.markdown("**Table 4: PP Strapping Cost**")
    
    # Initialize PP strapping costs in session state
    if 'pp_strapping_costs' not in st.session_state:
        st.session_state.pp_strapping_costs = pd.DataFrame({
            "Strapping Length/m": [1],
            "Cost (LKR/m)": [15.00]
        })
    
    # Create editable PP strapping costs table
    edited_pp_strapping_df = st.data_editor(
        st.session_state.pp_strapping_costs,
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "Strapping Length/m": st.column_config.NumberColumn("Strapping Length/m", required=True, min_value=0),
            "Cost (LKR/m)": st.column_config.NumberColumn("Cost (LKR/m)", required=True, min_value=0, format="%.2f")
        },
        key="pp_strapping_editor"
    )
    
    # Update session state
    st.session_state.pp_strapping_costs = edited_pp_strapping_df
    
    # Add Calculate button for secondary calculations
    st.markdown("---")
    calculate_secondary_col1, calculate_secondary_col2, calculate_secondary_col3 = st.columns([1, 1, 1])
    with calculate_secondary_col2:
        calculate_secondary_btn = st.button("🔢 Calculate Secondary Packing Costs", 
                                          type="primary", 
                                          use_container_width=True,
                                          on_click=trigger_secondary_calculation)
    
    # Display secondary calculations when triggered
    if st.session_state.calculate_secondary:
        # Section 1: Secondary Packing Cost Per Profile
        st.subheader("Total Secondary Packing Cost Per Profile")
        
        if not st.session_state.secondary_sku_data.empty:
            # Check which method is used
            use_method1 = not st.session_state.bundling_data.empty
            use_method2 = not st.session_state.bundle_size_data.empty
            
            if not use_method1 and not use_method2:
                st.warning("Please enter data in either Method 1 (Number of layers) or Method 2 (Size of bundle)")
            else:
                # Get data for calculations
                material_cost_dict = dict(zip(
                    st.session_state.secondary_material_costs["Material"],
                    st.session_state.secondary_material_costs["Cost/ m²"]
                ))
                box_data = st.session_state.secondary_box_costs.iloc[0]
                polybag_data = st.session_state.polybag_costs.iloc[0]
                stretchwrap_data = st.session_state.stretchwrap_costs.iloc[0]
                
                # Prepare calculations data
                secondary_calculations_data = []
                
                for _, sku in st.session_state.secondary_sku_data.iterrows():
                    try:
                        # Extract SKU dimensions
                        profile_width = float(sku["Width/mm"])
                        profile_height = float(sku["Height/mm"])
                        profile_length = float(sku["Length/mm"])
                        sku_no = sku["SKU No"]
                        
                        # Calculate Profiles per bundle based on selected method
                        profiles_per_bundle = 0
                        
                        if use_method1:
                            # Method 1: Number of layers method
                            bundling_data = st.session_state.bundling_data.iloc[0]
                            rows_per_bundle = float(bundling_data["Number of rows/bundle"])
                            layers_per_bundle = float(bundling_data["Number of layer/bundle"])
                            profiles_per_bundle = rows_per_bundle * layers_per_bundle
                            
                        elif use_method2:
                            # Method 2: Size of bundle method
                            bundle_size_data = st.session_state.bundle_size_data.iloc[0]
                            bundle_width = float(bundle_size_data["Bundle width/mm"])
                            bundle_height = float(bundle_size_data["Bundle Height/mm"])
                            
                            # Check for division by zero
                            if profile_width > 0 and profile_height > 0:
                                profiles_per_bundle = (bundle_width / profile_width) * (bundle_height / profile_height)
                        
                        # Calculate Packing cost(LKR/profile)
                        packing_cost_per_profile = 0
                        
                        if packing_type == "polybag":
                            # (Polybag Cost/(polybag size*24.5* profiles per bundle))*(profile length)
                            polybag_size_inches = polybag_data["Polybag size (inches)"]
                            polybag_cost_per_m = polybag_data["Cost/m (LKR/m)"]
                            
                            if polybag_size_inches > 0 and profiles_per_bundle > 0 and 24.5 > 0:
                                packing_cost_per_profile = (polybag_cost_per_m / 
                                                          (polybag_size_inches * 24.5 * profiles_per_bundle)) * profile_length
                        
                        elif packing_type == "cardboard box":
                            box_volume = box_data["Length(mm)"] * box_data["Width (mm)"] * box_data["Height (mm)"]
                            
                            # Get bundle dimensions based on method used
                            if use_method1:
                                bundling_data = st.session_state.bundling_data.iloc[0]
                                rows_per_bundle = float(bundling_data["Number of rows/bundle"])
                                layers_per_bundle = float(bundling_data["Number of layer/bundle"])
                                
                                width_prof_type = bundling_data["width prof.type"]
                                if width_prof_type == "W/mm":
                                    bundle_width = profile_width * rows_per_bundle
                                    bundle_height = profile_height * layers_per_bundle
                                else:
                                    bundle_width = profile_height * rows_per_bundle
                                    bundle_height = profile_width * layers_per_bundle
                                
                            else:
                                bundle_size_data = st.session_state.bundle_size_data.iloc[0]
                                bundle_width = float(bundle_size_data["Bundle width/mm"])
                                bundle_height = float(bundle_size_data["Bundle Height/mm"])
                            
                            if box_volume > 0 and profiles_per_bundle > 0:
                                packing_cost_per_profile = (box_data["Cost (LKR)"] / (box_volume * profiles_per_bundle)) * \
                                                          (bundle_height * bundle_width * profile_length)
                        
                        # Calculate Stretchwrap cost (LKR/prof.)
                        stretchwrap_cost_per_profile = 0
                        if eco_friendly_tab2 == "Stretch wrap":
                            if stretchwrap_data["Area (mm²)"] > 0 and profiles_per_bundle > 0:
                                # Get bundle dimensions based on method used
                                if use_method1:
                                    bundling_data = st.session_state.bundling_data.iloc[0]
                                    rows_per_bundle = float(bundling_data["Number of rows/bundle"])
                                    layers_per_bundle = float(bundling_data["Number of layer/bundle"])
                                    
                                    width_prof_type = bundling_data["width prof.type"]
                                    if width_prof_type == "W/mm":
                                        bundle_width = profile_width * rows_per_bundle
                                        bundle_height = profile_height * layers_per_bundle
                                    else:
                                        bundle_width = profile_height * rows_per_bundle
                                        bundle_height = profile_width * layers_per_bundle
                                        
                                else:
                                    bundle_size_data = st.session_state.bundle_size_data.iloc[0]
                                    bundle_width = float(bundle_size_data["Bundle width/mm"])
                                    bundle_height = float(bundle_size_data["Bundle Height/mm"])
                                
                                stretchwrap_cost_per_profile = (stretchwrap_data["Cost (LKR/mm²)"] / 
                                                               (stretchwrap_data["Area (mm²)"] * profiles_per_bundle)) * \
                                                              (bundle_width * bundle_height)
                        
                        # Calculate Protective tape cost (LKR/profile)
                        protective_tape_cost_per_profile = 0
                        if protective_tape_tab2 == "Yes":
                            # Get bundle dimensions for surface area calculation
                            if use_method1:
                                bundling_data = st.session_state.bundling_data.iloc[0]
                                rows_per_bundle = float(bundling_data["Number of rows/bundle"])
                                layers_per_bundle = float(bundling_data["Number of layer/bundle"])
                                
                                width_prof_type = bundling_data["width prof.type"]
                                if width_prof_type == "W/mm":
                                    bundle_width = profile_width * rows_per_bundle
                                    bundle_height = profile_height * layers_per_bundle
                                else:
                                    bundle_width = profile_height * rows_per_bundle
                                    bundle_height = profile_width * layers_per_bundle
                                    
                            else:
                                bundle_size_data = st.session_state.bundle_size_data.iloc[0]
                                bundle_width = float(bundle_size_data["Bundle width/mm"])
                                bundle_height = float(bundle_size_data["Bundle Height/mm"])
                            
                            # Calculate bundle surface area in m²
                            bundle_surface_area_m2 = (2 * ((bundle_width * bundle_height) + 
                                                         (bundle_width * profile_length) + 
                                                         (bundle_height * profile_length))) / (1000 * 1000)
                            
                            protective_tape_cost_per_m2 = material_cost_dict.get("Protective Tape", 0)
                            total_protective_tape_cost = bundle_surface_area_m2 * protective_tape_cost_per_m2
                            
                            if profiles_per_bundle > 0:
                                protective_tape_cost_per_profile = total_protective_tape_cost / profiles_per_bundle
                        
                        # Calculate Total cost per profile
                        total_cost_per_profile = (packing_cost_per_profile + 
                                                stretchwrap_cost_per_profile + 
                                                protective_tape_cost_per_profile)
                        
                        # Add to calculations data
                        secondary_calculations_data.append({
                            "SKU": sku_no,
                            "Profiles per bundle": round(profiles_per_bundle, 2),
                            "Packing type": packing_type,
                            "Packing cost(LKR/profile)": round(packing_cost_per_profile, 4),
                            "Stretchwrap cost (LKR/prof.)": round(stretchwrap_cost_per_profile, 4),
                            "Protective tape cost (LKR/profile)": round(protective_tape_cost_per_profile, 4),
                            "Total cost per profile": round(total_cost_per_profile, 4)
                        })
                        
                    except (ValueError, TypeError, ZeroDivisionError) as e:
                        continue
                
                if secondary_calculations_data:
                    st.session_state.secondary_calculations = pd.DataFrame(secondary_calculations_data)
                    st.dataframe(st.session_state.secondary_calculations, use_container_width=True)
                    
                    total_cost_all = sum(item["Total cost per profile"] for item in secondary_calculations_data)
                    st.metric("**Total Secondary Packing Cost (All SKUs)**", f"LKR {total_cost_all:,.4f}")
                else:
                    st.warning("Unable to calculate costs. Please check all input data is valid.")
        else:
            st.info("Enter SKU data to see secondary packing cost calculations.")
        
        # Section 2: Crate/Pallet Cost Calculations
        st.divider()
        st.markdown("**Total Crate/Pallet Cost Calculation**")
        
        if not st.session_state.secondary_sku_data.empty and not st.session_state.crate_pallet_data.empty:
            # Get data for calculations
            crate_data = st.session_state.crate_costs.iloc[0]
            pallet_data = st.session_state.pallet_costs.iloc[0]
            strapping_clip_data = st.session_state.strapping_clip_costs.iloc[0]
            pp_strapping_data = st.session_state.pp_strapping_costs.iloc[0]
            
            # Prepare calculations data
            crate_pallet_calculations_data = []
            
            # Create a dictionary to map SKU to its dimensions from SKU table
            sku_dimensions = {}
            for _, sku in st.session_state.secondary_sku_data.iterrows():
                try:
                    sku_dimensions[sku["SKU No"]] = {
                        "width": float(sku["Width/mm"]),
                        "height": float(sku["Height/mm"]),
                        "length": float(sku["Length/mm"])
                    }
                except (ValueError, TypeError):
                    continue
            
            # Process each entry in crate/pallet data
            for _, crate_pallet_row in st.session_state.crate_pallet_data.iterrows():
                try:
                    sku_no = crate_pallet_row["SKU"]
                    packing_method = crate_pallet_row["packing method"]
                    crate_pallet_width = float(crate_pallet_row["Width/mm"])
                    crate_pallet_height = float(crate_pallet_row["Height/mm"])
                    crate_pallet_length = float(crate_pallet_row["Length/mm"])
                    
                    # Get SKU dimensions
                    if sku_no not in sku_dimensions:
                        continue
                        
                    profile_width = sku_dimensions[sku_no]["width"]
                    profile_height = sku_dimensions[sku_no]["height"]
                    profile_length = sku_dimensions[sku_no]["length"]
                    
                    # Calculate profiles per pallet/crate
                    profiles_per_crate_pallet = 0
                    if profile_width > 0 and profile_height > 0:
                        profiles_per_crate_pallet = (crate_pallet_width / profile_width) * (crate_pallet_height / profile_height)
                    
                    # Calculate crate/pallet cost(LKR)
                    crate_pallet_cost = 0
                    if packing_method == "pallet":
                        # (((Crate/Pallet Dimensions Table width*height*length) / (Table 2: Pallet Cost width*height*length)) * Table 2: Pallet Cost cost)
                        crate_pallet_volume = crate_pallet_width * crate_pallet_height * crate_pallet_length
                        pallet_ref_volume = pallet_data["Pallet width/mm"] * pallet_data["Pallet Height/mm"] * crate_pallet_length
                        
                        if pallet_ref_volume > 0:
                            crate_pallet_cost = (crate_pallet_volume / pallet_ref_volume) * pallet_data["Cost (LKR)"]
                    
                    elif packing_method == "crate":
                        # (((Crate/Pallet Dimensions Table width*height*length) / (Table 1: Crate Cost width*height*length)) * Table 1: Crate Cost cost)
                        crate_pallet_volume = crate_pallet_width * crate_pallet_height * crate_pallet_length
                        crate_ref_volume = crate_data["Crate width/mm"] * crate_data["Crate Height/mm"] * crate_data["Crate Length/mm"]
                        
                        if crate_ref_volume > 0:
                            crate_pallet_cost = (crate_pallet_volume / crate_ref_volume) * crate_data["Cost (LKR)"]
                    
                    # Calculate packing cost per profile(LKR/prof)
                    packing_cost_per_profile = 0
                    if profiles_per_crate_pallet > 0:
                        packing_cost_per_profile = crate_pallet_cost / profiles_per_crate_pallet
                    
                    # Calculate Number of strapping clips
                    number_of_strapping_clips = 0
                    if crate_pallet_length > 0:
                        number_of_strapping_clips = np.ceil(crate_pallet_length / 500)
                    
                    # Calculate strapping clip cost per profile
                    strapping_clip_cost_per_profile = 0
                    if profiles_per_crate_pallet > 0:
                        total_clip_cost = number_of_strapping_clips * strapping_clip_data["Cost"]
                        strapping_clip_cost_per_profile = total_clip_cost / profiles_per_crate_pallet
                    
                    # Calculate PP strapping cost per profile
                    pp_strapping_cost_per_profile = 0
                    if profiles_per_crate_pallet > 0 and pp_strapping_data["Strapping Length/m"] > 0:
                        pp_strapping_cost_per_profile = (
                            ((crate_pallet_width + crate_pallet_height) / pp_strapping_data["Strapping Length/m"]) / 
                            profiles_per_crate_pallet
                        ) * (number_of_strapping_clips * 2) * pp_strapping_data["Cost (LKR/m)"]
                    
                    # Calculate PP strapping cost (total, not per profile - for display purposes)
                    pp_strapping_cost = pp_strapping_cost_per_profile * profiles_per_crate_pallet if profiles_per_crate_pallet > 0 else 0
                    
                    # Calculate Total cost
                    total_cost = packing_cost_per_profile + strapping_clip_cost_per_profile + pp_strapping_cost_per_profile
                    
                    # Add to calculations data
                    crate_pallet_calculations_data.append({
                        "SKU": sku_no,
                        "Packing method": packing_method,
                        "profiles per pallet/crate": round(profiles_per_crate_pallet, 2),
                        "crate/pallet cost(LKR)": round(crate_pallet_cost, 2),
                        "packing cost per profile(LKR/prof)": round(packing_cost_per_profile, 4),
                        "Number of strapping clips": int(number_of_strapping_clips),
                        "strapping clip cost per profile": round(strapping_clip_cost_per_profile, 4),
                        "PP strapping cost": round(pp_strapping_cost, 4),
                        "PP strapping cost per profile": round(pp_strapping_cost_per_profile, 4),
                        "Total cost": round(total_cost, 4)
                    })
                    
                except (ValueError, TypeError, ZeroDivisionError) as e:
                    continue
            
            if crate_pallet_calculations_data:
                st.session_state.crate_pallet_calculations = pd.DataFrame(crate_pallet_calculations_data)
                st.dataframe(st.session_state.crate_pallet_calculations, use_container_width=True)
                
                total_cost_all = sum(item["Total cost"] for item in crate_pallet_calculations_data)
                st.metric("**Total Crate/Pallet Cost (All SKUs)**", f"LKR {total_cost_all:,.4f}")
            else:
                st.warning("Unable to calculate crate/pallet costs. Please check all input data is valid.")
        else:
            st.info("Enter SKU data and crate/pallet dimensions to see crate/pallet cost calculations.")
        
        # Reset calculation flag after displaying
        st.session_state.calculate_secondary = False
    
    st.divider()
    
    # New Section: Special Comments under Secondary Packing
    st.subheader("Special Comments under Secondary Packing")
    
    # Determine protective tape comment based on finish selection
    protective_tape_comment = ""
    if finish_tab2 in ["PC", "WF", "Anodised"]:
        protective_tape_comment = "Protective tape required to avoid rejects"
    else:
        protective_tape_comment = "Protective tape is not mandatory"
    
    # Create the comments box
    comments_box = f"""
    **Packing Method Note:**
    
    1. Costing is done according to Secondary packing.
    
    2. The interleaving material is **"{eco_friendly_tab2}"**.
    
    3. {protective_tape_comment}
    
    4. Costing is inclusive of secondary packing - pallet or crate, however it is not inclusive of any labels or artwork. These will incur an additional charge.
    """
    
    st.info(comments_box)
